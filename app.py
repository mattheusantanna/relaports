import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO
from datetime import datetime, date

# =========================================================
# CONFIGURAÇÃO
# =========================================================

st.set_page_config(
    page_title="Relatórios Operacionais",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =========================================================
# CSS CUSTOMIZADO
# =========================================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

/* ── Reset & Base ── */
*, *::before, *::after { box-sizing: border-box; }

html, body, [data-testid="stAppViewContainer"] {
    font-family: 'DM Sans', sans-serif;
    background-color: #0f1117;
    color: #e8eaf0;
}

/* Esconde o texto em AMBOS os estados (aberto e fechado) */
[data-testid="stIconMaterial"] {
    display: none !important;
}

/* Adiciona ícone visual no botão de reabrir */
[data-testid="stSidebarCollapseButton"] button::after,
[data-testid="collapsedControl"] button::after {
    content: "<" !important;
    font-size: 1.4rem !important;
    color: #3a445a !important;
}

[data-testid="stSidebarCollapseButton"] button:hover::after,
[data-testid="collapsedControl"] button:hover::after {
    color: #e8eaf0 !important;
}

[data-testid="stAppViewContainer"] {
    background: #0f1117;
}

[data-testid="stMain"] {
    background: #0f1117;
}

/* Esconde o texto no botão de reabrir */
[data-testid="stExpandSidebarButton"] [data-testid="stIconMaterial"] {
    display: none !important;
}

/* Adiciona ícone visual no botão de reabrir */
[data-testid="stExpandSidebarButton"]::after {
    content: ">" !important;
    font-size: 1.6rem !important;
    color: #8892a4 !important;
}

[data-testid="stExpandSidebarButton"] {
    background: #1a2033 !important;
    border-radius: 0 8px 8px 0 !important;
    width: 24px !important;
    height: 48px !important;
    opacity: 1 !important;
}

[data-testid="stExpandSidebarButton"]:hover::after {
    color: #e8eaf0 !important;
}

[data-testid="stExpandSidebarButton"]:hover {
    background: #232d42 !important;
}

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: #161b27 !important;
    border-right: 1px solid #1e2535 !important;
}

[data-testid="stSidebar"] * {
    font-family: 'DM Sans', sans-serif !important;
}

[data-testid="stSidebar"] .stRadio label {
    color: #8892a4 !important;
    font-size: 0.88rem !important;
    font-weight: 500 !important;
    padding: 0.5rem 0 !important;
    transition: color 0.2s !important;
}

[data-testid="stSidebar"] .stRadio label:hover {
    color: #e8eaf0 !important;
}

/* ── Títulos ── */
h1 {
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 700 !important;
    font-size: 1.7rem !important;
    color: #f0f2f8 !important;
    letter-spacing: -0.02em !important;
    margin-bottom: 0.2rem !important;
}

h2, h3 {
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    color: #c8cedd !important;
    letter-spacing: -0.01em !important;
}

/* ── Inputs ── */
[data-testid="stTextInput"] label,
[data-testid="stTextArea"] label,
[data-testid="stDateInput"] label {
    font-size: 0.75rem !important;
    font-weight: 600 !important;
    color: #6b7a99 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.07em !important;
    margin-bottom: 4px !important;
}

[data-testid="stTextInput"] input,
[data-testid="stTextArea"] textarea,
[data-testid="stDateInput"] input {
    background: #1a2033 !important;
    border: 1px solid #232d42 !important;
    border-radius: 8px !important;
    color: #e8eaf0 !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.9rem !important;
    transition: border-color 0.2s, box-shadow 0.2s !important;
}

[data-testid="stTextInput"] input:focus,
[data-testid="stTextArea"] textarea:focus {
    border-color: #3d7fff !important;
    box-shadow: 0 0 0 3px rgba(61,127,255,0.12) !important;
    outline: none !important;
}

[data-testid="stTextInput"] input::placeholder,
[data-testid="stTextArea"] textarea::placeholder {
    color: #3a445a !important;
}

/* ── Botões ── */
[data-testid="stButton"] button {
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.85rem !important;
    border-radius: 8px !important;
    border: none !important;
    transition: all 0.2s !important;
    letter-spacing: 0.01em !important;
}

[data-testid="stButton"] button[kind="primary"],
[data-testid="stButton"] button {
    background: #1a2033 !important;
    color: #8892a4 !important;
    border: 1px solid #232d42 !important;
    padding: 0.45rem 1rem !important;
}

[data-testid="stButton"] button:hover {
    background: #232d42 !important;
    color: #e8eaf0 !important;
    border-color: #3d7fff !important;
    transform: translateY(-1px) !important;
}

/* Botão de gerar relatório – destaque */
.gerar-btn [data-testid="stButton"] button {
    background: linear-gradient(135deg, #3d7fff 0%, #1a5cdb 100%) !important;
    color: #fff !important;
    border: none !important;
    padding: 0.6rem 1.8rem !important;
    font-size: 0.92rem !important;
    box-shadow: 0 4px 20px rgba(61,127,255,0.3) !important;
}

.gerar-btn [data-testid="stButton"] button:hover {
    box-shadow: 0 6px 28px rgba(61,127,255,0.45) !important;
    transform: translateY(-2px) !important;
}

/* ── Download button ── */
[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #00c97a 0%, #009958 100%) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
    padding: 0.6rem 1.8rem !important;
    box-shadow: 0 4px 20px rgba(0,201,122,0.3) !important;
    transition: all 0.2s !important;
}

[data-testid="stDownloadButton"] button:hover {
    box-shadow: 0 6px 28px rgba(0,201,122,0.45) !important;
    transform: translateY(-2px) !important;
}

/* ── Divider ── */
hr {
    border-color: #1e2535 !important;
    margin: 1.5rem 0 !important;
}

/* ── Success / alerts ── */
[data-testid="stAlert"] {
    background: #0d2518 !important;
    border: 1px solid #1a5c3a !important;
    border-radius: 10px !important;
    color: #4ade80 !important;
}

/* ── Container com borda (bloco datas) ── */
[data-testid="stVerticalBlockBorderWrapper"] {
    background: #141928 !important;
    border: 1px solid #1e2a3e !important;
    border-radius: 12px !important;
    padding: 1rem !important;
}

/* ── Markdown labels de coluna ── */
.col-header p {
    font-size: 0.7rem !important;
    font-weight: 700 !important;
    color: #4a566e !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
    margin-bottom: 4px !important;
    padding-bottom: 6px !important;
    border-bottom: 1px solid #1e2535 !important;
}

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #0f1117; }
::-webkit-scrollbar-thumb { background: #232d42; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #3d7fff; }
</style>
""", unsafe_allow_html=True)


# =========================================================
# HELPERS
# =========================================================

def badge(text: str, color: str = "#3d7fff"):
    """Renderiza um badge colorido inline."""
    st.markdown(
        f'<span style="background:{color}22;color:{color};'
        f'font-size:0.7rem;font-weight:700;letter-spacing:0.06em;'
        f'text-transform:uppercase;padding:3px 10px;border-radius:20px;'
        f'border:1px solid {color}44;">{text}</span>',
        unsafe_allow_html=True
    )


def section_header(icon: str, title: str, subtitle: str = ""):
    st.markdown(
        f"""
        <div style="margin-bottom:1.2rem;">
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:4px;">
                <span style="font-size:1.3rem;">{icon}</span>
                <h1 style="margin:0!important;">{title}</h1>
            </div>
            {"<p style='color:#4a566e;font-size:0.88rem;margin:0 0 0 2rem;'>"+subtitle+"</p>" if subtitle else ""}
        </div>
        """,
        unsafe_allow_html=True
    )


def col_headers(headers: list):
    """Renderiza cabeçalhos de coluna estilizados."""
    cols = st.columns(len(headers))
    for col, h in zip(cols, headers):
        col.markdown(
            f'<div class="col-header"><p>{h}</p></div>',
            unsafe_allow_html=True
        )


def download_excel(wb, nome_arquivo):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    st.download_button(
        label="⬇️  Baixar Relatório Excel",
        data=output,
        file_name=nome_arquivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def linha_vazia_chegada():
    return {k: "" for k in [
        "data", "nota", "lote", "fardos",
        "peso_bruto", "peso_liquido", "tara", "placa",
        "peso_caminhao", "tara_caminhao",
        "peso_bruto_carga", "peso_liquido_carga"
    ]}


def linha_vazia_saida():
    return {k: "" for k in [
        "container", "nota", "lote", "fardos",
        "tara_cntr", "max_gross",
        "hora_inicio", "hora_fim", "data", "bruto_carga"
    ]}


def linha_vazia_estufagem():
    return {k: "" for k in [
        "nota_fiscal", "lote", "qtd_fardos", "peso", "obs"
    ]}


# =========================================================
# SIDEBAR
# =========================================================

with st.sidebar:
    st.markdown(
        """
        <div style="padding:1.2rem 0 1.8rem;">
            <div style="font-size:1.5rem;font-weight:700;color:#f0f2f8;
                        letter-spacing:-0.02em;margin-bottom:4px;">
                ⚓ Relatórios
            </div>
            <div style="font-size:0.78rem;color:#3d4a63;font-weight:500;">
                Sistema Operacional
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    relatorio = st.radio(
        "Módulo",
        [
            "📥  Chegada",
            "📤  Saída",
            "📦  Estufagem"
        ],
        label_visibility="collapsed"
    )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        '<div style="font-size:0.7rem;color:#2a3348;text-align:center;">'
        'v2.0 · Operações Portuárias</div>',
        unsafe_allow_html=True
    )


# =========================================================
# RELATÓRIO CHEGADA
# =========================================================

if "Chegada" in relatorio:

    TEMPLATE_PATH = "Chegada.xlsx"

    section_header("📥", "Peso Caminhão — Chegada",
                   "Registro de entrada e pesagem de carga")

    badge("Chegada", "#3d7fff")
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Campos de cabeçalho ──
    with st.container():
        c1, c2, c3 = st.columns([2, 2, 2])
        with c1:
            st.markdown("**Identificação**")
            produtor = st.text_input("Produtor", placeholder="Nome do produtor")
            terminal = st.text_input("Terminal", placeholder="Terminal de operação")
        with c2:
            st.markdown("**Documentação**")
            instrucao = st.text_input("Instrução", placeholder="Nº da instrução")
            data_relatorio = st.date_input("Data", datetime.today())
        with c3:
            st.markdown("**Responsável**")
            inspetor = st.text_input("Inspetor", placeholder="Nome do inspetor")

    observacao = st.text_area("Observação", height=90,
                              placeholder="Registre aqui observações relevantes...")

    st.divider()

    # ── Tabela de dados ──
    st.markdown("#### Registros de Pesagem")

    if "linhas_chegada" not in st.session_state:
        st.session_state.linhas_chegada = [linha_vazia_chegada()]

    headers_ch = [
        "Data", "Nota Fiscal", "Lote", "Qtd Fardos",
        "Data", "Nota Fiscal", "Lote", "Qtd Fardos",
        "Peso Bruto", "Peso Líquido", "Tara", "Placa",
        "Peso Caminhão", "Tara Caminhão",
        "Bruto Carga", "Líquido Carga"
    ]

    col_headers(headers_ch)

    for i, linha in enumerate(st.session_state.linhas_chegada):
        cols = st.columns(len(headers_ch))
        for idx, key in enumerate(linha.keys()):
            linha[key] = cols[idx].text_input(
                "", value=linha[key],
                key=f"chegada_{key}_{i}",
                label_visibility="collapsed"
            )

    c_add, c_spacer = st.columns([1, 5])
    with c_add:
        if st.button("＋  Adicionar linha", key="add_chegada"):
            st.session_state.linhas_chegada.append(linha_vazia_chegada())
            st.rerun()

    st.divider()

    # ── Gerar ──
    col_btn, _ = st.columns([2, 5])
    with col_btn:
        st.markdown('<div class="gerar-btn">', unsafe_allow_html=True)
        gerar = st.button("📥  Gerar Relatório", key="gerar_chegada", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    if gerar:
        try:
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb["Peso Caminhão - Chegada"]

            ws["B2"] = produtor
            ws["B3"] = terminal
            ws["F2"] = instrucao
            ws["F3"] = data_relatorio.strftime("%d/%m/%Y")
            ws["I2"] = inspetor
            ws["I2"].font = Font(name="Arial", size=12, bold=True, color="000000")
            ws["I2"].alignment = Alignment(horizontal="center", vertical="center")
            ws["B26"] = observacao
            ws["B26"].alignment = Alignment(wrap_text=True, vertical="top")

            for idx, linha in enumerate(st.session_state.linhas_chegada):
                row = 6 + idx
                vals = list(linha.values())
                for col_i, val in enumerate(vals, start=1):
                    ws.cell(row=row, column=col_i).value = val

            st.success("✅  Relatório gerado com sucesso!")
            download_excel(wb,
                f"chegada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        except FileNotFoundError:
            st.error("⚠️  Template 'Chegada.xlsx' não encontrado.")


# =========================================================
# RELATÓRIO SAÍDA
# =========================================================

elif "Saída" in relatorio:

    TEMPLATE_PATH = "Saida.xlsx"

    section_header("📤", "Peso Caminhão — Saída",
                   "Controle de saída e pesagem de containers")

    badge("Saída", "#f59e0b")
    st.markdown("<br>", unsafe_allow_html=True)

    c1, c2, c3 = st.columns([2, 2, 2])
    with c1:
        st.markdown("**Identificação**")
        produtor = st.text_input("Produtor", placeholder="Nome do produtor")
        terminal = st.text_input("Terminal", placeholder="Terminal de operação")
    with c2:
        st.markdown("**Documentação**")
        instrucao = st.text_input("Instrução", placeholder="Nº da instrução")
        data_relatorio = st.date_input("Data", datetime.today())
    with c3:
        st.markdown("**Limites de Peso**")
        limite_peso_cntr = st.text_input("Limite por Container",
                                         placeholder="kg")
        limite_total_instrucao = st.text_input("Limite Total da Instrução",
                                               placeholder="kg")

    observacao = st.text_area("Observação", height=90,
                              placeholder="Registre aqui observações relevantes...")

    st.divider()

    st.markdown("#### Registros de Saída")

    if "linhas_saida" not in st.session_state:
        st.session_state.linhas_saida = [linha_vazia_saida()]

    headers_sa = [
        "Container", "Nota Fiscal", "Lote", "Total Fardos",
        "Tara Cntr", "Max Gross",
        "Horário Início", "Horário Final", "Data", "Bruto Carga"
    ]

    col_headers(headers_sa)

    for i, linha in enumerate(st.session_state.linhas_saida):
        cols = st.columns(len(headers_sa))

        # Campos de texto simples: container, nota, lote, fardos, tara_cntr, max_gross, bruto_carga
        text_map = {
            0: "container",
            1: "nota",
            2: "lote",
            3: "fardos",
            4: "tara_cntr",
            5: "max_gross",
            9: "bruto_carga",
        }
        for col_idx, key in text_map.items():
            linha[key] = cols[col_idx].text_input(
                "", value=linha[key],
                key=f"saida_{key}_{i}",
                label_visibility="collapsed"
            )

        # Horário Início (coluna 6)
        try:
            hora_inicio_val = datetime.strptime(linha.get("hora_inicio", ""), "%H:%M").time()
        except (ValueError, TypeError, KeyError):
            hora_inicio_val = datetime.now().time().replace(second=0, microsecond=0)

        hora_inicio = cols[6].time_input(
            "",
            value=hora_inicio_val,
            key=f"saida_hora_inicio_{i}",
            label_visibility="collapsed"
        )
        linha["hora_inicio"] = hora_inicio.strftime("%H:%M")

        # Horário Final (coluna 7)
        try:
            hora_fim_val = datetime.strptime(linha.get("hora_fim", ""), "%H:%M").time()
        except (ValueError, TypeError, KeyError):
            hora_fim_val = datetime.now().time().replace(second=0, microsecond=0)

        hora_fim = cols[7].time_input(
            "",
            value=hora_fim_val,
            key=f"saida_hora_fim_{i}",
            label_visibility="collapsed"
        )
        linha["hora_fim"] = hora_fim.strftime("%H:%M")

        # Data (coluna 8)
        try:
            data_val = datetime.strptime(linha.get("data", ""), "%d/%m/%Y").date()
        except (ValueError, TypeError, KeyError):
            data_val = date.today()

        data_sel = cols[8].date_input(
            "",
            value=data_val,
            key=f"saida_data_{i}",
            label_visibility="collapsed"
        )
        linha["data"] = data_sel.strftime("%d/%m/%Y")

    c_add, _ = st.columns([1, 5])
    with c_add:
        if st.button("＋  Adicionar linha", key="add_saida"):
            st.session_state.linhas_saida.append(linha_vazia_saida())
            st.rerun()

    st.divider()

    col_btn, _ = st.columns([2, 5])
    with col_btn:
        st.markdown('<div class="gerar-btn">', unsafe_allow_html=True)
        gerar = st.button("📤  Gerar Relatório", key="gerar_saida", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    if gerar:
        try:
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb["Peso Caminhão - Saída"]

            ws["B2"] = produtor
            ws["B3"] = terminal
            ws["F2"] = instrucao
            ws["F3"] = data_relatorio.strftime("%d/%m/%Y")
            ws["I2"] = limite_peso_cntr
            ws["I3"] = limite_total_instrucao
            ws["B26"] = observacao
            ws["B26"].alignment = Alignment(wrap_text=True, vertical="top")

            for idx, linha in enumerate(st.session_state.linhas_saida):
                row = 5 + idx
                for col_i, val in enumerate(linha.values(), start=1):
                    ws.cell(row=row, column=col_i).value = val

            st.success("✅  Relatório gerado com sucesso!")
            download_excel(wb,
                f"saida_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        except FileNotFoundError:
            st.error("⚠️  Template 'Saida.xlsx' não encontrado.")


# =========================================================
# ESTUFAGEM
# =========================================================

elif "Estufagem" in relatorio:

    TEMPLATE_PATH = "Estufagem.xlsx"

    section_header("📦", "Estufagem Individual",
                   "Controle de estufagem e lacre de containers")

    badge("Estufagem", "#10b981")
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Campos de cabeçalho ──
    c1, c2, c3 = st.columns([2, 2, 2])
    with c1:
        st.markdown("**Identificação**")
        instrucao  = st.text_input("Instrução de Embarque", placeholder="Nº instrução")
        produtor   = st.text_input("Produtor", placeholder="Nome do produtor")
        container  = st.text_input("Nº Container", placeholder="Ex: TCKU1234567")
    with c2:
        st.markdown("**Dados Técnicos**")
        tara_porta = st.text_input("Tara Porta Cntr", placeholder="kg")
        max_gross  = st.text_input("Max Gross", placeholder="kg")
        lacre      = st.text_input("Lacre", placeholder="Nº do lacre")
    with c3:
        st.markdown("**Local & Responsável**")
        terminal   = st.text_input("Terminal", placeholder="Terminal de operação")
        qtd_fardos_total  = st.text_input("Qtd Total de Fardos", placeholder="Total")

    # ── Bloco datas/horários ──
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("**🕐 Datas e Horários da Estufagem**")

    with st.container(border=True):
        ca, cb, cc, cd = st.columns(4)
        with ca:
            hora_inicio = st.time_input("Começo Estufagem", value=datetime.now().time(), key="hora_inicio")
            inicio = hora_inicio.strftime("%H:%M")
        with cb:
            data_inicio = st.date_input("Data/Hora Início", datetime.today(), key="data_inicio")
            data_hora = data_inicio.strftime("%d/%m/%Y")
        with cc:
            hora_termino = st.time_input("Término Estufagem", value=datetime.now().time(), key="hora_termino")
            termino = hora_termino.strftime("%H:%M")
        with cd:
            data_termino = st.date_input("Data/Hora Término", datetime.today(), key="data_termino")
            data_hora_termino = data_termino.strftime("%d/%m/%Y")

    st.divider()

    # ── Tabela ──
    st.markdown("#### 📦 Dados da Estufagem")

    if "linhas_estufagem" not in st.session_state:
        st.session_state.linhas_estufagem = [linha_vazia_estufagem()]

    headers_es = ["Nota Fiscal", "Lote", "Qtd Fardos", "Peso", "Obs."]
    col_headers(headers_es)

    for i, linha in enumerate(st.session_state.linhas_estufagem):
        cols = st.columns(len(headers_es))
        for idx, key in enumerate(linha.keys()):
            linha[key] = cols[idx].text_input(
                "", value=linha[key],
                key=f"estufagem_{key}_{i}",
                label_visibility="collapsed"
            )

    c_add, _ = st.columns([1, 5])
    with c_add:
        if st.button("＋  Adicionar linha", key="add_estufagem"):
            st.session_state.linhas_estufagem.append(linha_vazia_estufagem())
            st.rerun()

    st.divider()

    # ── Inspetor + Observação ──
    c_ins, _ = st.columns([2, 4])
    with c_ins:
        inspetor = st.text_input("Inspetor", placeholder="Nome do inspetor responsável")

    observacao = st.text_area("Observação Final", height=120,
                              placeholder="Registre aqui observações finais sobre a estufagem...")

    st.markdown("<br>", unsafe_allow_html=True)

    col_btn, _ = st.columns([2, 5])
    with col_btn:
        st.markdown('<div class="gerar-btn">', unsafe_allow_html=True)
        gerar = st.button("📦  Gerar Relatório", key="gerar_estufagem", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    if gerar:
        try:
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb["Estufagem individual"]

            ws["B2"] = instrucao
            ws["B3"] = produtor
            ws["B4"] = container
            ws["B5"] = tara_porta
            ws["D5"] = max_gross
            ws["D6"] = lacre
            ws["B6"] = terminal
            ws["D4"] = qtd_fardos_total
            ws["B7"] = inicio
            ws["D7"] = data_hora
            ws["B8"] = termino
            ws["D8"] = data_hora_termino
            ws["A29"] = inspetor

            for idx, linha in enumerate(st.session_state.linhas_estufagem):
                row = 11 + idx
                for col_i, val in enumerate(linha.values(), start=1):
                    ws.cell(row=row, column=col_i).value = val

            ws["B26"] = observacao
            ws["B26"].alignment = Alignment(wrap_text=True, vertical="top")

            st.success("✅  Relatório gerado com sucesso!")
            download_excel(wb,
                f"estufagem_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        except FileNotFoundError:
            st.error("⚠️  Template 'Estufagem.xlsx' não encontrado.")
